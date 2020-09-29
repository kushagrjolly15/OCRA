from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
from PIL import Image
import os
import io
import urllib
import pytesseract
import pandas as pd
import os
from openpyxl import Workbook,load_workbook
from datetime import date, datetime
import glob
import time
import requests
import msal
import atexit
import os.path
import re
import base64
import json
from pdf2image import convert_from_path, convert_from_bytes
from pascal_voc_writer import Writer as Writer_P_voc
# import jinja2
# from jinja2 import Environment, PackageLoader
# import cv2
from shutil import copyfile
import sys
import dropbox
from dropbox.files import WriteMode
from dropbox.exceptions import ApiError, AuthError

import logging

# Access token
DROPBOX_TOKEN = 'RPzQlg0RBMoAAAAAAAAAAWrS6yRpmEXxmS9FPZmVChOzI4pdYlpKzKl6Oij_iUY5'


app = Flask(__name__)

logging.basicConfig(filename='demo.log',level=logging.DEBUG)


@app.route('/', methods=['POST','GET'])
def index():
    global writer
    if (len(DROPBOX_TOKEN) == 0):
        sys.exit("ERROR: Looks like you didn't add your access token. Open up backup-and-restore-example.py in a text editor and paste in your token in line 14.")

    # Create an instance of a Dropbox class, which can make requests to the API.
    print("Creating a Dropbox object...")
    dbx = dropbox.Dropbox(DROPBOX_TOKEN)

    # Check that the access token is valid
    try:
        dbx.users_get_current_account()
    except AuthError as err:
        sys.exit(
            "ERROR: Invalid access token; try re-generating an access token from the app console on the web.")
    directory = []
    try:
        for entry in dbx.files_list_folder('/ECopy-LRs/').entries:
            if (isinstance(entry, dropbox.files.FolderMetadata)):
                directory.append(entry.name)
    except Error as err:
        sys.exit("Error while checking file details")
    metadata, res = dbx.files_download(path="/ECopy-LRs/.OCRA_VM/Master_userlist.xlsx")
    df = pd.read_excel(io.BytesIO(res.content))
    user_list = df['Users'].tolist()
    if request.method == 'POST':
        if 'pull_directory' in request.json:
            pulled_directory = request.json['pull_directory']
            sub_directory = []
            try:
                for entry in dbx.files_list_folder('/ECopy-LRs/'+pulled_directory).entries:
                    if (isinstance(entry, dropbox.files.FolderMetadata)):
                        sub_directory.append(entry.name)
            except Error as err:
                sys.exit("Error while checking file details")
            return {'sub_directory': sub_directory}
        elif 'split' in request.json:
            if request.json['split'] == 'Split':
                encoded_string, page_no, total_pages = split(request.json['file'])
                return {'page': encoded_string, 'page_no' : page_no, 'total_pages' : total_pages}
        elif 'page_number' in request.json:
            encoded_string, page_no = get_page(request.json['page_number'])
            return {'page': encoded_string, 'page_no' : page_no}
        elif 'button_name' in request.json:
            if request.json['button_name'] == 'Submit':
                print('here 1')
                print(request.json['path'].split('/')[2].split('.')[0])
                file = request.json['file']
                directory = request.json['directory']
                sub_directory = request.json['sub_directory']
                title = request.json['title']
                journal = request.json['journal']
                volume = request.json['volumn']
                issue = request.json['issue']
                page = request.json['page']
                mode = request.json['mode']
                user = request.json['user']
                date_of_publish = datetime.strptime(request.json['date'], "%Y-%m-%d").strftime('%m.%d.%Y')
                date_of_publish_excel = datetime.strptime(request.json['date'], "%Y-%m-%d").strftime('%m/%d/%Y')
                # print(date_of_publish)
                # directory = re.sub('[^A-Za-z0-9? ]+', '', directory)
                # sub_directory = re.sub('[^A-Za-z0-9? ]+', '', sub_directory)
                # title = re.sub('[^A-Za-z0-9? ]+', '', title)
                # journal = re.sub('[^A-Za-z0-9? ]+', '', journal)
                # volume = re.sub('[^A-Za-z0-9? ]+', '', volume)
                # issue = re.sub('[^A-Za-z0-9? ]+', '', issue)
                # page = re.sub('[^A-Za-z0-9? ]+', '', page)
                if not journal:
                    journal = " "
                if not volume:
                    volume = " "
                if not issue:
                    issue = " "
                if not page:
                    page = " "
                if not mode:
                    mode = " "
                if not user:
                    user = " "
                # current_day = date.today()
                # formatted_date = date.strftime(current_day, "%m.%d.%Y")
                file = file.split(',')[1]
                file_name = directory.strip() + "_" + sub_directory.strip() + "_" + title.strip() + "_" + journal.strip() + "_" + volume.strip() + "_" + issue.strip() + "_" + page.strip() + "_" + date_of_publish.strip() + "_" + mode.strip() + "_" + user.strip() + ".pdf"
                file_path = '/static/tmp_download/'
                for f in os.listdir(os.getcwd() + file_path):
                    os.remove(os.getcwd() + file_path + f)
                print('here2')
                writer.save(os.getcwd() + '/static/tmp_download/' + request.json['path'].split('/')[2].split('.')[0] + '.xml')
                print(writer)
                with open(os.getcwd() + '/static/tmp_download/' + request.json['path'].split('/')[2].split('.')[0] + '.xml', 'rb') as f:
                    # We use WriteMode=overwrite to make sure that the settings in the file
                    # are changed on upload
                    print("Uploading " + os.getcwd() + '/static/tmp_download/' + request.json['path'].split('/')[2].split('.')[0] + '.xml' + " to Dropbox as " + '/ECopy-LRs/.OCRA/OCRA_ODM_Dataset/Annotations/')
                    try:
                        dbx.files_upload(f.read(), '/ECopy-LRs/.OCRA/OCRA_ODM_Dataset/Annotations/' + request.json['path'].split('/')[2].split('.')[0] + '.xml' , mode=WriteMode('overwrite'))
                    except ApiError as err:
                        # This checks for the specific error where a user doesn't have enough Dropbox space quota to upload this file
                        if (err.error.is_path() and
                                err.error.get_path().error.is_insufficient_space()):
                            sys.exit("ERROR: Cannot back up; insufficient space.")
                        elif err.user_message_text:
                            print(err.user_message_text)
                            sys.exit()
                        else:
                            print(err)
                            sys.exit()
                with open(os.getcwd() + '/static/' + request.json['path'], 'rb') as f:
                    # We use WriteMode=overwrite to make sure that the settings in the file
                    # are changed on upload
                    print("Uploading " + os.getcwd() + '/static/' + request.json['path'] + " to Dropbox as " + '/ECopy-LRs/.OCRA/OCRA_ODM_Dataset/Images/')
                    try:
                        dbx.files_upload(f.read(), '/ECopy-LRs/.OCRA/OCRA_ODM_Dataset/Images/' + request.json['path'].split('/')[2], mode=WriteMode('overwrite'))
                    except ApiError as err:
                        # This checks for the specific error where a user doesn't have enough Dropbox space quota to upload this file
                        if (err.error.is_path() and
                                err.error.get_path().error.is_insufficient_space()):
                            sys.exit("ERROR: Cannot back up; insufficient space.")
                        elif err.user_message_text:
                            print(err.user_message_text)
                            sys.exit()
                        else:
                            print(err)
                            sys.exit()
                with open(os.getcwd() + file_path + file_name, 'wb+') as theFile:
                    theFile.write(base64.b64decode(file))
                with open(os.getcwd() + file_path + file_name, 'rb') as f:
                    # We use WriteMode=overwrite to make sure that the settings in the file
                    # are changed on upload
                    print("Uploading " + os.getcwd() + file_path + file_name + " to Dropbox as " + '/ECopy-LRs/'+ directory + '/' + sub_directory+ '/' + file_name + "...")
                    try:
                        dbx.files_upload(f.read(), '/ECopy-LRs/'+ directory + '/' + sub_directory+ '/' + file_name, mode=WriteMode('overwrite'))
                    except ApiError as err:
                        # This checks for the specific error where a user doesn't have enough Dropbox space quota to upload this file
                        if (err.error.is_path() and
                                err.error.get_path().error.is_insufficient_space()):
                            sys.exit("ERROR: Cannot back up; insufficient space.")
                        elif err.user_message_text:
                            print(err.user_message_text)
                            sys.exit()
                        else:
                            print(err)
                            sys.exit()
                # if directory in os.listdir('/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'):
                #     if sub_directory in os.listdir('/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory):
                #         copyfile(os.getcwd() + file_path + file_name, '/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory+'/'+sub_directory+'/'+file_name)
                #     else:
                #         os.mkdir('/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory+'/'+sub_directory)
                #         copyfile(os.getcwd() + file_path + file_name, '/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory+'/'+sub_directory+'/'+file_name)
                # else:
                #     os.mkdir('/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory)
                #     if sub_directory in os.listdir('/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory):
                #         copyfile(os.getcwd() + file_path + file_name, '/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory+'/'+sub_directory+'/'+file_name)
                #     else:
                #         os.mkdir('/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory+'/'+sub_directory)
                #         copyfile(os.getcwd() + file_path + file_name, '/Users/kushagrjolly/Dropbox (ASU)/ECopy-LRs/'+directory+'/'+sub_directory+'/'+file_name)
                date_entered = date.today().strftime('%m/%d/%Y')
                if os.path.isfile(os.getcwd()+'/static/data/final_data.xlsx'):
                    print(os.getcwd())
                    workbook = load_workbook(filename=os.getcwd()+'/static/data/final_data.xlsx')
                    sheet = workbook.active
                    row = (directory,sub_directory,title,journal,volume,issue,page,date_of_publish_excel,mode,user,date_entered)
                    sheet.append(row)
                    workbook.save(filename=os.getcwd()+"/static/data/final_data.xlsx")
                else:
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.append(["Directory", "Subdirectory", "Title", "Journal/Wsite", "Volume", "Issue", "Pages", "Published_Date", "Category", "Entered by", "Date Entered"])
                    row = (directory, sub_directory, title, journal, volume, issue, page, date_of_publish_excel, mode, user, date_entered )
                    sheet.append(row)
                    workbook.save(filename=os.getcwd()+"/static/data/final_data.xlsx")
                print('here 3')
            else:
                data = request.json['path']
               	annotation_file = data + '.xml'
                file_path = "/static/"
                writer = Writer_P_voc(os.getcwd()+file_path+data, request.json['width'], request.json['height'], segmented=1, database = 'ODM')
                coordinates = request.json['coordinates']
                dim = (request.json['width'], request.json['height'])
                # img = cv2.imread(os.getcwd()+file_path+data)
                # img = cv2.resize(img, dim, interpolation = cv2.INTER_AREA)
                # cv2.imwrite(os.getcwd()+file_path+'new.jpeg', img)
                # crop_img = img[coordinates['y'] : coordinates['y'] + coordinates['h'], coordinates['x']: coordinates['x'] + coordinates['w']]
                # cv2.imwrite(os.getcwd()+'/static/img/cropped_test.jpg', crop_img)
                #
                img = Image.open(os.getcwd()+file_path+data)
                img = img.resize((request.json['width'], request.json['height']))
                img.save(os.getcwd()+file_path+"new.jpeg")
                width, height = img.size
                print(coordinates['x']+coordinates['w'])
                print(coordinates['y']+coordinates['h'])

                img2 = img.crop((coordinates['x'], coordinates['y'], coordinates['x']+coordinates['w'], coordinates['y']+coordinates['h']))
                crop_img = img2.convert('RGB')
                crop_img.save(os.getcwd()+'/static/img/cropped_test.jpg')
                text = pytesseract.image_to_string(crop_img)
                text = text.replace("\n", " ")
                text = re.sub(r'[^a-zA-Z0-9 ]+', '', text)
                writer.addObject(request.json['button_name'], coordinates['x'], coordinates['y'], coordinates['x']+coordinates['w'], coordinates['y']+coordinates['h'])
                if request.json['button_name'] == "Title:":
                    title = text
                    return {'title': title}
                elif request.json['button_name'] == "Journal:":
                    journal = text
                    return {'journal': journal}
                elif request.json['button_name'] == "Code:":
                    code = text
                    return {'code': code}
                elif request.json['button_name'] == "Volume:":
                    volume = text
                    return {'volume': volume}
                elif request.json['button_name'] == "Issue:":
                    issue = text
                    return {'issue': issue}
                elif request.json['button_name'] == "Page:":
                    page = text
                    return {'page': page}
		
    return render_template('index.html', directories = directory, user_list = user_list)


@app.route('/upload', methods=['POST'])
def upload():
    if (len(DROPBOX_TOKEN) == 0):
        sys.exit("ERROR: Looks like you didn't add your access token. Open up backup-and-restore-example.py in a text editor and paste in your token in line 14.")

    # Create an instance of a Dropbox class, which can make requests to the API.
    print("Creating a Dropbox object...")
    dbx = dropbox.Dropbox(DROPBOX_TOKEN)

    # Check that the access token is valid
    try:
        dbx.users_get_current_account()
    except AuthError as err:
        sys.exit(
            "ERROR: Invalid access token; try re-generating an access token from the app console on the web.")
    if request.json['upload']:
        if os.path.isfile('/home/kjolly5/OCRA/static/data/final_data.xlsx'):
                workbook = load_workbook(filename="/home/kjolly5/OCRA/static/data/final_data.xlsx")
                sheet = workbook.active
                rows = sheet.max_row
                columns = sheet.max_column
                data = []
                for row in sheet.iter_rows(min_row=2, min_col=1, max_row=rows, max_col=columns):
                    data_rows = []
                    for cell in row:
                        data_rows.append(cell.value)
                    data.append(data_rows)
                tuples = (tuple(x) for x in data)
                try:
                    dbx.files_download_to_file('/home/kjolly5/OCRA/static/data/Analytics - CAS Research.xlsx', '/ECopy-LRs/.OCRA_VM/Analytics - CAS Research.xlsx')
                except ApiError as err:
                    # This checks for the specific error where a user doesn't have enough Dropbox space quota to upload this file
                    if (err.error.is_path() and
                            err.error.get_path().error.is_insufficient_space()):
                        sys.exit("ERROR: Cannot download up")
                    elif err.user_message_text:
                        print(err.user_message_text)
                        sys.exit()
                    else:
                        print(err)
                        sys.exit()
                workbook1 = load_workbook(filename="/home/kjolly5/OCRA/static/data/Analytics - CAS Research.xlsx")
                sheet1 = workbook1["Dbase GP Jornals-Web"]
                for tup in tuples:
                    sheet1.append(tup)
                workbook1.save('/home/kjolly5/OCRA/static/data/Analytics - CAS Research.xlsx')
                with open('/home/kjolly5/OCRA/static/data/Analytics - CAS Research.xlsx', 'rb') as f:
                    # We use WriteMode=overwrite to make sure that the settings in the file
                    # are changed on upload
                    print("Uploading /home/kjolly5/OCRA/static/data/Analytics - CAS Research.xlsx to Dropbox as /ECopy-LRs/.OCRA_VM/Analytics - CAS Research.xlsx ...")
                    try:
                        dbx.files_upload(f.read(), '/ECopy-LRs/.OCRA_VM/Analytics - CAS Research.xlsx', mode=WriteMode('overwrite'))
                    except ApiError as err:
                        # This checks for the specific error where a user doesn't have enough Dropbox space quota to upload this file
                        if (err.error.is_path() and
                                err.error.get_path().error.is_insufficient_space()):
                            sys.exit("ERROR: Cannot back up; insufficient space.")
                        elif err.user_message_text:
                            print(err.user_message_text)
                            sys.exit()
                        else:
                            print(err)
                            sys.exit()
        for f in os.listdir('/home/kjolly5/OCRA/static/data/'):
            os.remove('/home/kjolly5/OCRA/static/data/' + f)
        # TENANT_ID = '41f88ecb-ca63-404d-97dd-ab0a169fd138'
        # CLIENT_ID = 'e46a9e18-13d4-4c0b-acd8-68450a62316b'
        # SHAREPOINT_HOST_NAME = 'arizonastateu-my.sharepoint.com'
        # SITE_NAME = 'personal'
        #
        #
        # AUTHORITY = 'https://login.microsoftonline.com/41f88ecb-ca63-404d-97dd-ab0a169fd138/'
        # ENDPOINT = 'https://graph.microsoft.com/v1.0'
        #
        # SCOPES = [
        #     'Files.ReadWrite.All',
        #     'Sites.ReadWrite.All',
        #     'User.Read',
        #     'User.ReadBasic.All'
        # ]
        #
        # cache = msal.SerializableTokenCache()
        #
        # if os.path.exists('token_cache.bin'):
        #     cache.deserialize(open('token_cache.bin', 'r').read())
        #
        # atexit.register(lambda: open('token_cache.bin', 'w').write(cache.serialize()) if cache.has_state_changed else None)
        #
        # app = msal.PublicClientApplication(CLIENT_ID, authority=AUTHORITY, token_cache=cache)
        #
        # accounts = app.get_accounts()
        # result = None
        # if len(accounts) > 0:
        #     result = app.acquire_token_silent(SCOPES, account=accounts[0])
        #
        # if result is None:
        #     flow = app.initiate_device_flow(scopes=SCOPES)
        #     if 'user_code' not in flow:
        #         raise Exception('Failed to create device flow')
        #
        #     print(flow['message'])
        #
        #     result = app.acquire_token_by_device_flow(flow)
        # if 'access_token' in result:
        #     access_token = result['access_token']
        #     headers={'Authorization': 'Bearer ' + access_token}
        #
        #     # result = requests.get(f'{ENDPOINT}/users/scalder1@asurite.asu.edu/drive/root/children', headers=headers)
        #     result = requests.get(f'{ENDPOINT}/me/drive/root/children/Analytics - CAS Research.xlsx/content', headers=headers)
        #     # filename="/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/Analytics - CAS Research.xlsx"
        #     result.raise_for_status()
        #     open('/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/Analytics - CAS Research.xlsx', 'wb').write(result.content)
        #     if os.path.isfile('/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/final_data.xlsx'):
        #         workbook = load_workbook(filename="/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/final_data.xlsx")
        #         sheet = workbook.active
        #         rows = sheet.max_row
        #         columns = sheet.max_column
        #         data = []
        #         for row in sheet.iter_rows(min_row=2, min_col=1, max_row=rows, max_col=columns):
        #             data_rows = []
        #             for cell in row:
        #                 data_rows.append(cell.value)
        #             data.append(data_rows)
        #         tuples = (tuple(x) for x in data)
        #         # test = ('Canvacc', None, 'Delivery of mRNA vaccines with heterocyclic lipids increases anti-tumor efficacy by STING-mediated immune cell activation', 'Nature Biotechnology', 37, None, '1174-1185', datetime.datetime(2019, 10, 1, 0, 0), 'J', 'AK', datetime.datetime(2019, 11, 6, 0, 0))
        #         workbook1 = load_workbook(filename="/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/Analytics - CAS Research.xlsx")
        #         sheet1 = workbook1["Dbase GP Jornals-Web"]
        #         for tup in tuples:
        #             sheet1.append(tup)
        #         workbook1.save('/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/Analytics - CAS Research.xlsx')
        #         result = requests.put(
        #             f'{ENDPOINT}/me/drive/root/children/Analytics - CAS Research.xlsx/content',
        #             headers={
        #                 'Authorization': 'Bearer ' + access_token,
        #                 'Content-type': 'application/binary'
        #             },
        #             data=open("/Users/kushagrjolly/Desktop/OnCampus/CAS/OCR/OCR/static/data/Analytics - CAS Research.xlsx", 'rb').read()
        #         )
        # else:
        #     raise Exception('no access token in result')
    return redirect('/')

def split(decode):
    app.logger.info('Processing default request')
    total_pages = 0
    encoded_string = ''
    file_path = "/static/tmp/"
    page_no = 1
    file = decode
    file = file.split(',')[1]
    file_name = 'temp.pdf'
    for f in os.listdir('/home/kjolly5/OCRA/static/tmp/'):
        os.remove('/home/kjolly5/OCRA/static/tmp/' + f)
    with open('/home/kjolly5/OCRA/static/tmp/'+ file_name, 'wb+') as theFile:
        theFile.write(base64.b64decode(file))
    pages = convert_from_path('/home/kjolly5/OCRA/static/tmp/' + file_name , 300, output_folder = '/home/kjolly5/OCRA/static/tmp/', fmt='jpeg')
    app.logger.info('Processing default request 2')
    for filename in os.listdir('/home/kjolly5/OCRA/static/tmp/'):
        if filename.endswith(".jpg"):
            total_pages += 1
            page_number = int(filename.split('.')[0].split('-')[-1])
            if page_no == page_number:
                encoded_string = '/tmp/'+filename
                # with open(os.getcwd() + file_path + filename, "rb") as image_file:
                #     encoded_string = json.dumps(base64.b64encode(image_file.read()))
        else:
            continue

    return (encoded_string, page_no, total_pages)

def get_page(page_no):
    file_path = "/static/tmp/"
    for filename in os.listdir('/home/kjolly5/OCRA/static/tmp/'):
        if filename.endswith(".jpg"):
            page_number = int(filename.split('.')[0].split('-')[-1])
            if page_no == page_number:
                encoded_string = '/tmp/'+filename
    return (encoded_string, page_no)

if __name__ == "__main__":
    app.run(host='0.0.0.0')
